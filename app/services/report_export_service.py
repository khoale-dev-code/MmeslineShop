"""Configurable, branded Excel export for Analytics Intelligence."""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.models.report_models import ExportSelection


FOREST = "1B4922"
FOREST_DARK = "123418"
GOLD = "C99E14"
CREAM = "FBFAF4"
WHITE = "FFFFFF"
INK = "101510"
MUTED = "667263"
PALE_GREEN = "EAF1E9"
PALE_GOLD = "F7EECF"
PALE_RED = "FCE8E8"
LINE = "DDE5DB"


class ReportExportService:
    CHANNEL_LABELS = {
        "website": "Website",
        "pos": "Cửa hàng (POS)",
        "shopee": "Shopee",
        "lazada": "Lazada",
        "tiktok_shop": "TikTok Shop",
        "tiktok": "TikTok Shop",
    }
    SEGMENT_LABELS = {
        "stock_risk": "Rủi ro thiếu hàng",
        "accelerating": "Đang tăng tốc",
        "bestseller": "Bán chạy",
        "potential": "Có tiềm năng",
        "slow": "Bán chậm",
        "stable": "Ổn định",
    }
    FORECAST_STATUS_LABELS = {
        "ready": "Đủ dữ liệu",
        "insufficient": "Chưa đủ dữ liệu",
        "disabled": "Chưa bật",
    }
    SHEETS = {
        "summary": "Tổng quan điều hành",
        "products": "Hiệu suất sản phẩm",
        "channels": "So sánh kênh",
        "forecast": "Dự báo & nhập hàng",
        "quality": "Chất lượng dữ liệu",
    }
    PRODUCT_COLUMNS = {
        "sku": ("SKU", "text"),
        "name": ("Sản phẩm", "text"),
        "segment": ("Phân nhóm", "text"),
        "sold_units": ("Đã bán", "integer"),
        "previous_units": ("Kỳ trước", "integer"),
        "net_revenue": ("Doanh thu thuần", "currency"),
        "growth_pct": ("Tăng trưởng", "percent_points"),
        "gross_profit": ("Lợi nhuận gộp", "currency"),
        "gross_margin": ("Biên lợi nhuận", "percent_points"),
        "views": ("Lượt xem", "integer"),
        "carts": ("Thêm giỏ", "integer"),
        "wishlists": ("Wishlist", "integer"),
        "conversion": ("Chuyển đổi", "percent_points"),
        "stock": ("Tồn kho", "integer"),
        "forecast_30d": ("Dự báo 30 ngày", "integer"),
        "forecast_low": ("Dự báo thấp", "integer"),
        "forecast_high": ("Dự báo cao", "integer"),
        "reorder_qty": ("Nên nhập thêm", "integer"),
        "opportunity_score": ("Điểm cơ hội", "decimal"),
        "forecast_confidence": ("Độ tin cậy", "percent_points"),
        "reasons": ("Giải thích", "text"),
    }
    DEFAULT_PRODUCT_COLUMNS = (
        "sku",
        "name",
        "segment",
        "sold_units",
        "net_revenue",
        "growth_pct",
        "gross_margin",
        "stock",
        "forecast_30d",
        "reorder_qty",
        "opportunity_score",
        "forecast_confidence",
        "reasons",
    )

    @classmethod
    def normalize_selection(cls, payload: dict[str, Any] | None) -> ExportSelection:
        payload = payload or {}
        sheets = tuple(
            key for key in payload.get("sheets", ())
            if key in cls.SHEETS
        ) or tuple(cls.SHEETS)
        product_columns = tuple(
            key for key in payload.get("product_columns", ())
            if key in cls.PRODUCT_COLUMNS
        ) or cls.DEFAULT_PRODUCT_COLUMNS
        title = str(payload.get("title") or "GUAMAISON Analytics Intelligence").strip()[:120]
        note = str(payload.get("note") or "").strip()[:500]
        return ExportSelection(
            sheets=sheets,
            product_columns=product_columns,
            include_charts=bool(payload.get("include_charts", True)),
            title=title,
            note=note,
        )

    @classmethod
    def preview(cls, report: dict[str, Any], selection: ExportSelection) -> dict[str, Any]:
        counts = {
            "summary": len(report.get("kpis") or {}),
            "products": len(report.get("products") or []),
            "channels": len(report.get("channels") or []),
            "forecast": sum(
                1 for row in (report.get("products") or [])
                if row.get("forecast_30d") is not None
            ),
            "quality": len((report.get("data_quality") or {}).get("issues") or []),
        }
        return {
            "sheets": [
                {"key": key, "label": cls.SHEETS[key], "row_count": counts[key]}
                for key in selection.sheets
            ],
            "product_columns": [
                {"key": key, "label": cls.PRODUCT_COLUMNS[key][0]}
                for key in selection.product_columns
            ],
            "include_charts": selection.include_charts,
        }

    def build(self, report: dict[str, Any], selection: ExportSelection) -> BytesIO:
        workbook = Workbook()
        workbook.remove(workbook.active)
        workbook.properties.title = selection.title
        workbook.properties.subject = "GUAMAISON omnichannel product analytics"
        workbook.properties.creator = "GUAMAISON Analytics Intelligence"

        data_sheet = workbook.create_sheet("_Data")
        self._write_chart_data(data_sheet, report)

        for key in selection.sheets:
            if key == "summary":
                self._write_summary(workbook, report, selection)
            elif key == "products":
                self._write_products(workbook, report, selection)
            elif key == "channels":
                self._write_channels(workbook, report)
            elif key == "forecast":
                self._write_forecast(workbook, report)
            elif key == "quality":
                self._write_quality(workbook, report)

        data_sheet.sheet_state = "hidden"
        workbook.active = 1 if len(workbook.worksheets) > 1 else 0
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    def _write_summary(
        self,
        workbook: Workbook,
        report: dict[str, Any],
        selection: ExportSelection,
    ) -> None:
        sheet = workbook.create_sheet(self.SHEETS["summary"])
        self._base_sheet(sheet, landscape=True)
        sheet.merge_cells("A1:H2")
        title = sheet["A1"]
        title.value = selection.title
        title.font = Font(name="Aptos Display", size=24, bold=True, color=WHITE)
        title.fill = PatternFill("solid", fgColor=FOREST_DARK)
        title.alignment = Alignment(vertical="center", horizontal="left")

        metadata = report.get("metadata") or {}
        filters = metadata.get("filters") or {}
        sheet.merge_cells("A3:H3")
        sheet["A3"] = (
            f"Kỳ báo cáo: {filters.get('start_date', '')} → {filters.get('end_date', '')}"
            f"  |  Tạo lúc: {str(metadata.get('generated_at') or '')[:19]} UTC"
        )
        sheet["A3"].font = Font(color=MUTED, italic=True, size=10)
        sheet["A3"].alignment = Alignment(vertical="center")

        if selection.note:
            sheet.merge_cells("A4:H4")
            sheet["A4"] = selection.note
            sheet["A4"].fill = PatternFill("solid", fgColor=PALE_GOLD)
            sheet["A4"].alignment = Alignment(wrap_text=True, vertical="center")

        kpis = report.get("kpis") or {}
        cards = [
            ("Doanh thu thuần", (kpis.get("net_revenue") or {}).get("value"), "currency"),
            ("Số đơn hợp lệ", (kpis.get("orders") or {}).get("value"), "integer"),
            ("Sản phẩm đã bán", (kpis.get("sold_units") or {}).get("value"), "integer"),
            ("Giá trị đơn TB", (kpis.get("average_order_value") or {}).get("value"), "currency"),
            ("Biên lợi nhuận", (kpis.get("gross_margin") or {}).get("value"), "percent_points"),
            ("Tỷ lệ chuyển đổi", (kpis.get("conversion") or {}).get("value"), "percent_points"),
            ("Chất lượng dữ liệu", (kpis.get("data_quality") or {}).get("value"), "percent_points"),
            ("Dự báo 30 ngày", (report.get("forecast") or {}).get("revenue_30d"), "currency"),
        ]
        for index, (label, value, kind) in enumerate(cards):
            column = (index % 4) * 2 + 1
            row = 6 + (index // 4) * 3
            sheet.merge_cells(start_row=row, start_column=column, end_row=row, end_column=column + 1)
            sheet.merge_cells(start_row=row + 1, start_column=column, end_row=row + 2, end_column=column + 1)
            label_cell = sheet.cell(row=row, column=column, value=label)
            value_cell = sheet.cell(row=row + 1, column=column, value=value if value is not None else "N/A")
            for target in (label_cell, value_cell):
                target.fill = PatternFill("solid", fgColor=CREAM)
                target.border = Border(bottom=Side(style="thin", color=LINE))
            label_cell.font = Font(size=9, bold=True, color=MUTED)
            value_cell.font = Font(size=18, bold=True, color=FOREST_DARK)
            value_cell.alignment = Alignment(vertical="center")
            self._apply_number_format(value_cell, kind)

        if selection.include_charts:
            self._add_summary_charts(sheet, report)

        quality = report.get("data_quality") or {}
        sheet.merge_cells("A29:H29")
        sheet["A29"] = f"MODEL STATUS: {'PASS' if quality.get('score', 0) >= 70 else 'CHECK'}"
        sheet["A29"].font = Font(bold=True, color=WHITE)
        sheet["A29"].fill = PatternFill(
            "solid",
            fgColor=FOREST if quality.get("score", 0) >= 70 else GOLD,
        )
        sheet["A30"] = "Nguồn"
        sheet["B30"] = "Website / POS / Marketplace API đã cấp quyền"
        sheet["A31"] = "Phạm vi"
        sheet["B31"] = (metadata.get("data_scope") or "")
        sheet.merge_cells("B30:H30")
        sheet.merge_cells("B31:H31")
        sheet.freeze_panes = "A5"

    def _write_products(
        self,
        workbook: Workbook,
        report: dict[str, Any],
        selection: ExportSelection,
    ) -> None:
        sheet = workbook.create_sheet(self.SHEETS["products"])
        columns = list(selection.product_columns)
        headers = [self.PRODUCT_COLUMNS[key][0] for key in columns]
        rows = []
        for product in report.get("products") or []:
            row = []
            for key in columns:
                value = product.get(key)
                if key == "reasons":
                    value = " • ".join(value or [])
                elif key == "segment":
                    value = self.SEGMENT_LABELS.get(str(value or ""), value)
                row.append(value)
            rows.append(row)
        self._write_table_sheet(sheet, "ProductPerformanceTable", headers, rows)

        for col_index, key in enumerate(columns, start=1):
            kind = self.PRODUCT_COLUMNS[key][1]
            for row_index in range(2, len(rows) + 2):
                cell = sheet.cell(row=row_index, column=col_index)
                self._apply_number_format(cell, kind)
                if key == "reasons":
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        for key in ("growth_pct", "gross_margin", "conversion", "opportunity_score", "forecast_confidence"):
            if key in columns and rows:
                column = get_column_letter(columns.index(key) + 1)
                sheet.conditional_formatting.add(
                    f"{column}2:{column}{len(rows) + 1}",
                    ColorScaleRule(
                        start_type="min", start_color="F8C7C7",
                        mid_type="percentile", mid_value=50, mid_color="F7EECF",
                        end_type="max", end_color="BFD8BE",
                    ),
                )
        if "reorder_qty" in columns and rows:
            column = get_column_letter(columns.index("reorder_qty") + 1)
            sheet.conditional_formatting.add(
                f"{column}2:{column}{len(rows) + 1}",
                CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor=PALE_RED)),
            )

    def _write_channels(self, workbook: Workbook, report: dict[str, Any]) -> None:
        sheet = workbook.create_sheet(self.SHEETS["channels"])
        headers = ["Kênh", "Doanh thu thuần", "Tỷ trọng", "Tăng trưởng", "Đã bán", "Chuyển đổi", "Tracking"]
        rows = [[
            self.CHANNEL_LABELS.get(str(row.get("channel") or ""), row.get("channel")),
            row.get("net_revenue"),
            row.get("revenue_share"),
            row.get("growth_pct"),
            row.get("sold_units"),
            row.get("conversion"),
            "Đủ" if row.get("tracking_complete") else "N/A",
        ] for row in (report.get("channels") or [])]
        self._write_table_sheet(sheet, "ChannelComparisonTable", headers, rows)
        for row in range(2, len(rows) + 2):
            self._apply_number_format(sheet.cell(row, 2), "currency")
            for column in (3, 4, 6):
                self._apply_number_format(sheet.cell(row, column), "percent_points")

    def _write_forecast(self, workbook: Workbook, report: dict[str, Any]) -> None:
        sheet = workbook.create_sheet(self.SHEETS["forecast"])
        headers = [
            "SKU", "Sản phẩm", "Tồn kho", "Dự báo 30 ngày", "Thấp", "Cao",
            "Nên nhập thêm", "Độ tin cậy", "Trạng thái", "Giải thích",
        ]
        rows = [[
            row.get("sku"), row.get("name"), row.get("stock"), row.get("forecast_30d"),
            row.get("forecast_low"), row.get("forecast_high"), row.get("reorder_qty"),
            row.get("forecast_confidence"),
            self.FORECAST_STATUS_LABELS.get(
                str(row.get("forecast_status") or ""),
                row.get("forecast_status"),
            ),
            " • ".join(row.get("reasons") or []),
        ] for row in (report.get("products") or [])]
        self._write_table_sheet(sheet, "ForecastReorderTable", headers, rows)
        for row in range(2, len(rows) + 2):
            self._apply_number_format(sheet.cell(row, 8), "percent_points")
            sheet.cell(row, 10).alignment = Alignment(wrap_text=True, vertical="top")
        if rows:
            sheet.conditional_formatting.add(
                f"G2:G{len(rows) + 1}",
                CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor=PALE_RED)),
            )

    def _write_quality(self, workbook: Workbook, report: dict[str, Any]) -> None:
        sheet = workbook.create_sheet(self.SHEETS["quality"])
        quality = report.get("data_quality") or {}
        headers = ["Kiểm tra", "Kết quả", "Trạng thái / nơi cần xử lý"]
        rows = [
            ["Điểm chất lượng", quality.get("score"), quality.get("label")],
            ["Độ phủ giá vốn", quality.get("cost_coverage"), "Cần ≥ 80% để tin cậy lợi nhuận"],
            ["Dòng tracking", quality.get("tracking_rows"), "product_analytics"],
            ["Đơn trong kỳ", quality.get("order_rows"), "orders / order_items"],
            ["Sản phẩm phân tích", quality.get("product_rows"), "products"],
        ]
        rows.extend([[f"Cảnh báo {index}", None, issue] for index, issue in enumerate(quality.get("issues") or [], 1)])
        self._write_table_sheet(sheet, "DataQualityTable", headers, rows)
        self._apply_number_format(sheet["B2"], "percent_points")
        self._apply_number_format(sheet["B3"], "percent_points")
        for row in range(2, len(rows) + 2):
            sheet.cell(row, 3).alignment = Alignment(wrap_text=True, vertical="top")

    def _write_chart_data(self, sheet, report: dict[str, Any]) -> None:
        sheet["A1"] = "Kỳ"
        sheet["B1"] = "Doanh thu thực tế"
        for index, point in enumerate((report.get("trend") or {}).get("points") or [], start=2):
            sheet.cell(index, 1, point.get("label"))
            sheet.cell(index, 2, point.get("revenue"))

        sheet["D1"] = "Kênh"
        sheet["E1"] = "Doanh thu"
        for index, row in enumerate(report.get("channels") or [], start=2):
            channel = str(row.get("channel") or "")
            sheet.cell(index, 4, self.CHANNEL_LABELS.get(channel, channel))
            sheet.cell(index, 5, row.get("net_revenue"))

        sheet["G1"] = "Dự báo"
        sheet["H1"] = "Thấp"
        sheet["I1"] = "Cơ sở"
        sheet["J1"] = "Cao"
        for index, row in enumerate((report.get("forecast") or {}).get("points") or [], start=2):
            sheet.cell(index, 7, row.get("label"))
            sheet.cell(index, 8, row.get("low"))
            sheet.cell(index, 9, row.get("revenue"))
            sheet.cell(index, 10, row.get("high"))

    def _add_summary_charts(self, sheet, report: dict[str, Any]) -> None:
        trend_count = len((report.get("trend") or {}).get("points") or [])
        if trend_count:
            chart = LineChart()
            chart.title = "Xu hướng doanh thu thuần (VNĐ)"
            chart.y_axis.title = "VNĐ"
            chart.style = 13
            chart.height = 8
            chart.width = 14
            chart.legend = None
            chart.add_data(Reference(sheet.parent["_Data"], min_col=2, min_row=1, max_row=trend_count + 1), titles_from_data=True)
            chart.set_categories(Reference(sheet.parent["_Data"], min_col=1, min_row=2, max_row=trend_count + 1))
            sheet.add_chart(chart, "A13")

        channel_count = len(report.get("channels") or [])
        if channel_count:
            chart = BarChart()
            chart.title = "Doanh thu theo kênh (VNĐ)"
            chart.style = 10
            chart.height = 8
            chart.width = 12
            chart.legend = None
            chart.add_data(Reference(sheet.parent["_Data"], min_col=5, min_row=1, max_row=channel_count + 1), titles_from_data=True)
            chart.set_categories(Reference(sheet.parent["_Data"], min_col=4, min_row=2, max_row=channel_count + 1))
            sheet.add_chart(chart, "E13")

    def _write_table_sheet(self, sheet, table_name: str, headers: list[str], rows: list[list[Any]]) -> None:
        self._base_sheet(sheet, landscape=True)
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        last_row = max(1, len(rows) + 1)
        last_col = len(headers)
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor=FOREST_DARK)
            cell.font = Font(bold=True, color=WHITE)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.row_dimensions[1].height = 32
        if rows:
            table = Table(displayName=table_name, ref=f"A1:{get_column_letter(last_col)}{last_row}")
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium4",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(table)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"
        self._fit_columns(sheet, max_width=46)

    @staticmethod
    def _base_sheet(sheet, *, landscape: bool) -> None:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A2"
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.orientation = "landscape" if landscape else "portrait"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.outlinePr.summaryBelow = True
        sheet.oddFooter.center.text = "GUAMAISON Analytics Intelligence"
        sheet.oddFooter.right.text = "Trang &P / &N"
        for column in range(1, 9):
            sheet.column_dimensions[get_column_letter(column)].width = 16

    @staticmethod
    def _fit_columns(sheet, max_width: int = 40) -> None:
        for column_cells in sheet.columns:
            letter = get_column_letter(column_cells[0].column)
            width = max((len(str(cell.value or "")) for cell in column_cells), default=8) + 2
            sheet.column_dimensions[letter].width = max(10, min(max_width, width))

    @staticmethod
    def _apply_number_format(cell, kind: str) -> None:
        if cell.value is None or isinstance(cell.value, str):
            return
        if kind == "currency":
            cell.number_format = '#,##0 "₫";[Red](#,##0) "₫";-'
        elif kind == "integer":
            cell.number_format = '#,##0;[Red](#,##0);-'
        elif kind == "decimal":
            cell.number_format = '0.0'
        elif kind == "percent_points":
            cell.value = float(cell.value) / 100
            cell.number_format = '0.0%'
