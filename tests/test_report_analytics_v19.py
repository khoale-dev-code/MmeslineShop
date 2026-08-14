from __future__ import annotations

import unittest
from datetime import date

from openpyxl import load_workbook

from app.models.report_models import ReportFilters, ReportSnapshot
from app.services.report_analytics_service import ReportAnalyticsService
from app.services.report_export_service import ReportExportService


class FakeRepository:
    def __init__(self, snapshot: ReportSnapshot) -> None:
        self.snapshot = snapshot

    def load_snapshot(self, _filters: ReportFilters) -> ReportSnapshot:
        return self.snapshot


def filters() -> ReportFilters:
    return ReportFilters(date(2026, 7, 1), date(2026, 7, 30))


class ReportAnalyticsV19Tests(unittest.TestCase):
    def test_actual_units_are_not_duplicated_per_analytics_row(self) -> None:
        snapshot = ReportSnapshot(
            orders=[{
                "id": "o1", "total_amount": 500_000, "shipping_fee": 0,
                "sales_channel": "web", "created_at": "2026-07-10T08:00:00Z",
                "status": "completed", "payment_status": "paid",
            }],
            order_items=[{
                "order_id": "o1", "product_id": "p1", "quantity": 5,
                "unit_price": 100_000,
            }],
            products=[{
                "id": "p1", "name": "Áo sơ mi", "sku": "A01", "stock": 20,
                "price": 100_000, "cost_price": 60_000,
            }],
            analytics=[
                {"product_id": "p1", "channel": "web", "views": 100, "add_to_carts": 20, "created_at": "2026-07-08"},
                {"product_id": "p1", "channel": "web", "views": 100, "add_to_carts": 20, "created_at": "2026-07-09"},
            ],
        )
        report = ReportAnalyticsService(FakeRepository(snapshot)).build(filters())
        self.assertEqual(report["funnel"][0]["sold"], 5)
        self.assertEqual(report["funnel"][0]["views"], 200)
        self.assertEqual(report["funnel"][0]["conversion"], 2.5)

    def test_missing_tracking_is_na_not_a_fake_hundred_percent(self) -> None:
        snapshot = ReportSnapshot(
            orders=[{
                "id": "o1", "total_amount": 100_000, "sales_channel": "pos",
                "created_at": "2026-07-10", "status": "completed", "payment_status": "paid",
            }],
            order_items=[{"order_id": "o1", "product_id": "p1", "quantity": 1, "unit_price": 100_000}],
            products=[{"id": "p1", "name": "Quần", "stock": 3, "price": 100_000, "cost_price": 50_000}],
        )
        report = ReportAnalyticsService(FakeRepository(snapshot)).build(filters())
        funnel = report["funnel"][0]
        self.assertIsNone(funnel["views"])
        self.assertIsNone(funnel["conversion"])
        self.assertFalse(funnel["tracking_complete"])

    def test_forecast_stays_off_when_history_is_too_short(self) -> None:
        snapshot = ReportSnapshot(
            orders=[{
                "id": "o1", "total_amount": 200_000, "sales_channel": "web",
                "created_at": "2026-07-20", "status": "completed", "payment_status": "paid",
            }],
            order_items=[{"order_id": "o1", "product_id": "p1", "quantity": 2, "unit_price": 100_000}],
            products=[{"id": "p1", "name": "Váy", "stock": 8, "price": 100_000, "cost_price": 55_000}],
        )
        report = ReportAnalyticsService(FakeRepository(snapshot)).build(filters())
        self.assertIsNone(report["products"][0]["forecast_30d"])
        self.assertEqual(report["forecast"]["status"], "insufficient")

    def test_connected_marketplace_uses_normalized_external_orders(self) -> None:
        snapshot = ReportSnapshot(
            products=[{"id": "p1", "name": "Áo khoác", "sku": "AK01", "stock": 9, "price": 250_000, "cost_price": 120_000}],
            marketplace_connections=[{"provider": "shopee", "status": "active", "shop_id": "shop-1"}],
            external_orders=[{
                "id": "eo1", "provider": "shopee", "shop_id": "shop-1",
                "order_status": "completed", "payment_status": "paid",
                "net_amount": 500_000, "ordered_at": "2026-07-14T09:00:00Z",
            }],
            external_order_items=[{
                "external_order_pk": "eo1", "external_product_id": "sp1",
                "external_sku_id": "sku1", "quantity": 2, "returned_quantity": 0,
                "unit_price": 250_000, "item_discount": 0,
            }],
            product_channel_mappings=[{
                "provider": "shopee", "shop_id": "shop-1", "external_product_id": "sp1",
                "external_sku_id": "sku1", "product_id": "p1",
            }],
        )
        report = ReportAnalyticsService(FakeRepository(snapshot)).build(filters())
        self.assertEqual(report["channels"][0]["channel"], "shopee")
        self.assertEqual(report["channels"][0]["net_revenue"], 500_000)
        self.assertEqual(report["products"][0]["sold_units"], 2)

    def test_excel_export_contains_selected_sheets(self) -> None:
        snapshot = ReportSnapshot()
        report = ReportAnalyticsService(FakeRepository(snapshot)).build(filters())
        exporter = ReportExportService()
        selection = exporter.normalize_selection({
            "sheets": ["summary", "quality"],
            "product_columns": ["sku", "name"],
            "include_charts": False,
        })
        buffer = exporter.build(report, selection)
        workbook = load_workbook(buffer, data_only=False)
        self.assertIn("Tổng quan điều hành", workbook.sheetnames)
        self.assertIn("Chất lượng dữ liệu", workbook.sheetnames)
        self.assertNotIn("Hiệu suất sản phẩm", workbook.sheetnames)
        self.assertEqual(workbook["_Data"].sheet_state, "hidden")


if __name__ == "__main__":
    unittest.main()
